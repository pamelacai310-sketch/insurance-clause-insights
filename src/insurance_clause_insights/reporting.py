from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .config import FIELD_DISPLAY_NAMES
from .models import ComparisonGroup


def _display_key_facts(key_facts: dict[str, str]) -> dict[str, str]:
    return {FIELD_DISPLAY_NAMES.get(key, key): value for key, value in key_facts.items()}


def serialize_groups(groups: list[ComparisonGroup], category_counts: dict[str, int]) -> dict:
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "category_counts": category_counts,
        "groups": [
            {
                "category": group.category,
                "product_count": group.product_count,
                "products": [
                    {
                        "company": product.company,
                        "product_name": product.product_name,
                        "pdf_path": product.pdf_path,
                        "source_url": product.source_url,
                        "key_facts": _display_key_facts(product.key_facts),
                        "unique_features": [
                            {
                                "label": feature.label,
                                "snippet": feature.snippet,
                                "score": feature.score,
                            }
                            for feature in product.unique_features
                        ],
                        # 新增：精算参数
                        "actuarial_params": {
                            "entry_age": product.entry_age,
                            "gender": product.gender,
                            "annual_premium": product.annual_premium,
                            "sum_assured": product.sum_assured,
                            "payment_period": product.payment_period,
                            "insurance_period": product.insurance_period,
                            "dividend_type": product.dividend_type,
                            "guaranteed_rate": product.guaranteed_rate,
                        },
                    }
                    for product in group.products
                ],
            }
            for group in groups
        ],
    }


def write_json_report(groups: list[ComparisonGroup], category_counts: dict[str, int], output_path: Path) -> None:
    output_path.write_text(
        json.dumps(serialize_groups(groups, category_counts), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_markdown_report(groups: list[ComparisonGroup], category_counts: dict[str, int], output_path: Path) -> None:
    lines: list[str] = [
        "# 保险条款横向比较报告",
        "",
        f"- 生成时间: {datetime.now().isoformat(timespec='seconds')}",
        f"- 类别数量概览: {', '.join(f'{name}:{count}' for name, count in sorted(category_counts.items(), key=lambda item: (-item[1], item[0])))}",
        "",
    ]

    for group in groups:
        lines.extend(
            [
                f"## {group.category}（{group.product_count} 款）",
                "",
            ]
        )
        for product in group.products:
            lines.append(f"### {product.company} - {product.product_name}")
            if product.key_facts:
                display_facts = _display_key_facts(product.key_facts)
                lines.append(
                    f"- 关键字段: {'；'.join(f'{key}={value}' for key, value in display_facts.items())}"
                )
            lines.append(f"- 合同 PDF: `{product.pdf_path}`")
            for index, feature in enumerate(product.unique_features, start=1):
                lines.append(
                    f"- 特色{index} [{feature.label}]（score={feature.score:.4f}）: {feature.snippet}"
                )
            lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")


def write_excel_report(groups: list[ComparisonGroup], category_counts: dict[str, int], output_path: Path) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "产品特色"

    headers = [
        "类别",
        "公司",
        "产品名",
        "保险类型",
        "保险期间",
        "缴费期间",
        "等待期",
        "合同PDF",
        "特色1",
        "特色2",
        "特色3",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col_index, header in enumerate(headers, start=1):
        cell = overview.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_index = 2
    for group in groups:
        for product in group.products:
            feature_texts = [feature.snippet for feature in product.unique_features[:3]]
            while len(feature_texts) < 3:
                feature_texts.append("")
            overview.append(
                [
                    group.category,
                    product.company,
                    product.product_name,
                    product.key_facts.get("insurance_type", ""),
                    product.key_facts.get("insurance_period", ""),
                    product.key_facts.get("payment_period", ""),
                    product.key_facts.get("waiting_period", ""),
                    product.pdf_path,
                    feature_texts[0],
                    feature_texts[1],
                    feature_texts[2],
                ]
            )
            for cell in overview[row_index]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row_index += 1

    for width, column in zip((16, 12, 30, 20, 18, 18, 14, 48, 40, 40, 40), "ABCDEFGHIJK"):
        overview.column_dimensions[column].width = width

    stats = workbook.create_sheet("类别统计")
    stats.append(["类别", "条款数量"])
    for name, count in sorted(category_counts.items(), key=lambda item: (-item[1], item[0])):
        stats.append([name, count])

    workbook.save(output_path)
